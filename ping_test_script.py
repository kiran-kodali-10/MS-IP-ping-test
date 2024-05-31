# import pandas as pd
# import os
# import platform
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# from datetime import datetime
# import signal
# import sys

# # Global flag to handle interruption
# interrupted = False

# def signal_handler(sig, frame):
#     global interrupted
#     print("Interrupt signal received. Preparing to exit...")
#     interrupted = True

# # Register the signal handler for SIGINT (Ctrl+C)
# signal.signal(signal.SIGINT, signal_handler)

# def ping_ip(ip):
#     # Determine the command based on the OS
#     param = '-n' if platform.system().lower() == 'windows' else '-c'
#     command = ['ping', param, '1', ip]
#     response = os.system(" ".join(command))
#     return response == 0

# def update_ip_status(file_path):
#     # Load the existing Excel file
#     wb = load_workbook(file_path)
#     sheet = wb.active
    
#     # Read the data into a DataFrame
#     df = pd.read_excel(file_path)
    
#     # Ensure there is a column named 'expected IP' (case insensitive) for IP addresses
#     ip_column = next((col for col in df.columns if col.lower() == 'expected ip'), None)
#     if not ip_column:
#         raise ValueError("Excel file must contain a column named 'expected IP'")
    
#     # Ensure 'Status' and 'Last Updated' columns exist
#     if 'Status' not in df.columns:
#         df['Status'] = ''
#     if 'Last Updated' not in df.columns:
#         df['Last Updated'] = ''
    
#     # Save the initial DataFrame if new columns were added
#     df.to_excel(file_path, index=False)
    
#     # Define fill colors for Reachable and Not Reachable statuses
#     reachable_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
#     not_reachable_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    
#     try:
#         # Check and update the status of each IP address
#         for index, row in df.iterrows():
#             global interrupted
#             if interrupted:
#                 print("Interrupted by user. Saving progress and exiting.")
#                 break

#             ip = row[ip_column]
#             if pd.isna(ip) or ip == '':
#                 continue  # Skip if IP address is missing or empty
            
#             ip = str(ip)
#             status = 'Reachable' if ping_ip(ip) else 'Not Reachable'
            
#             # Get the current timestamp
#             timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
#             # Determine the fill color based on status
#             fill = reachable_fill if status == 'Reachable' else not_reachable_fill
            
#             # Update the status and timestamp in the Excel file directly
#             status_cell = sheet.cell(row=index+2, column=df.columns.get_loc('Status')+1)
#             timestamp_cell = sheet.cell(row=index+2, column=df.columns.get_loc('Last Updated')+1)
            
#             status_cell.value = status
#             status_cell.fill = fill
#             timestamp_cell.value = timestamp
#             timestamp_cell.fill = fill
            
#             # Save the workbook after each update
#             wb.save(file_path)
#     except Exception as e:
#         print(f"An error occurred: {e}")
#     finally:
#         # Ensure the workbook is saved before exiting
#         wb.save(file_path)
#         print("Progress saved. Exiting.")

# if __name__ == "__main__":
#     if len(sys.argv) != 2:
#         print("Usage: python script.py <path_to_excel_file>")
#         sys.exit(1)
    
#     # Specify the path to your Excel file
#     file_path = sys.argv[1]

#     # Add 'Status' and 'Last Updated' columns if they don't exist
#     df = pd.read_excel(file_path)
#     if 'Status' not in df.columns or 'Last Updated' not in df.columns:
#         if 'Status' not in df.columns:
#             df['Status'] = ''
#         if 'Last Updated' not in df.columns:
#             df['Last Updated'] = ''
#         df.to_excel(file_path, index=False)

#     # Update the IP status in the Excel file
#     update_ip_status(file_path)

import pandas as pd
import os
import platform
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import signal
import sys

# Custom exception to handle keyboard interrupt
class KeyboardInterruptError(Exception):
    pass

def signal_handler(sig, frame):
    print("Interrupt signal received. Preparing to exit...")
    raise KeyboardInterruptError

# Register the signal handler for SIGINT (Ctrl+C)
signal.signal(signal.SIGINT, signal_handler)

def ping_ip(ip):
    # Determine the command based on the OS
    param = '-n' if platform.system().lower() == 'windows' else '-c'
    command = ['ping', param, '1', ip]
    response = os.system(" ".join(command))
    return response == 0

def update_ip_status(file_path):
    # Load the existing Excel file
    wb = load_workbook(file_path)
    sheet = wb.active
    
    # Read the data into a DataFrame
    df = pd.read_excel(file_path)
    
    # Ensure there is a column named 'expected IP' (case insensitive) for IP addresses
    ip_column = next((col for col in df.columns if col.lower() == 'expected ip'), None)
    if not ip_column:
        raise ValueError("Excel file must contain a column named 'expected IP'")
    
    # Ensure 'Status' and 'Last Updated' columns exist
    if 'Status' not in df.columns:
        df['Status'] = ''
    if 'Last Updated' not in df.columns:
        df['Last Updated'] = ''
    
    # Save the initial DataFrame if new columns were added
    df.to_excel(file_path, index=False)
    
    # Define fill colors for Reachable and Not Reachable statuses
    reachable_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    not_reachable_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    
    try:
        # Check and update the status of each IP address
        for index, row in df.iterrows():
            ip = row[ip_column]
            if pd.isna(ip) or ip == '':
                continue  # Skip if IP address is missing or empty
            
            ip = str(ip)
            status = 'Reachable' if ping_ip(ip) else 'Not Reachable'
            
            # Get the current timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Determine the fill color based on status
            fill = reachable_fill if status == 'Reachable' else not_reachable_fill
            
            # Update the status and timestamp in the Excel file directly
            status_cell = sheet.cell(row=index+2, column=df.columns.get_loc('Status')+1)
            timestamp_cell = sheet.cell(row=index+2, column=df.columns.get_loc('Last Updated')+1)
            
            status_cell.value = status
            status_cell.fill = fill
            timestamp_cell.value = timestamp
            timestamp_cell.fill = fill
            
            # Save the workbook after each update
            wb.save(file_path)
    except KeyboardInterruptError:
        print("Keyboard interrupt received. Saving progress and exiting.")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Ensure the workbook is saved before exiting
        wb.save(file_path)
        print("Progress saved. Exiting.")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <path_to_excel_file>")
        sys.exit(1)
    
    # Specify the path to your Excel file
    file_path = sys.argv[1]

    # Add 'Status' and 'Last Updated' columns if they don't exist
    df = pd.read_excel(file_path)
    if 'Status' not in df.columns or 'Last Updated' not in df.columns:
        if 'Status' not in df.columns:
            df['Status'] = ''
        if 'Last Updated' not in df.columns:
            df['Last Updated'] = ''
        df.to_excel(file_path, index=False)

    # Update the IP status in the Excel file
    update_ip_status(file_path)
